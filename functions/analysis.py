# library imports
import os
import re
from datetime import datetime
from collections import Counter

# Openpyxl library imports
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList


#### XSLX FUNCTIONS ####
# apply style for standard sheet, procedure
def apply_style(sheet):
    for cell in sheet[1]: # first row
        cell.font = Font(bold=True)
    sheet.freeze_panes = 'A2' # header freeze pane
    sheet.auto_filter.ref = sheet.dimensions # add filters

# write tabular data into a xslx sheet, procedure
def write_sheet(sheet, dataset, flag_matches=False):
    if dataset != []:
        if flag_matches:
            headers = ['IR_id', 'OA_id', 'Confidence', 'matched_fields', 'matched_length']
            sheet.append(headers)
            for row1 in dataset:
                fields_string = '; '.join(row1[3])
                row1[3] = fields_string
                sheet.append(row1)
        else:
            actual_keys = list(dataset[0].keys())
            headers = actual_keys[:]  # create a copy of actual_keys and change fields
            idx = headers.index('authors')
            headers[idx:idx+1] = ['authors_names', 'authors_orcids']

            sheet.append(headers)
            for row2 in dataset:
                row2_values = []
                for head in actual_keys:
                    if head == 'authors':
                        authors_names = '; '.join( f"{aut['surname']}, {aut['name']}" for aut in row2['authors'] )
                        authors_orcids = '; '.join(aut2['orcid'] for aut2 in row2['authors'])
                        row2_values.append(authors_names)
                        row2_values.append(authors_orcids)
                    elif head == 'issn':
                        issn_string = '; '.join(iss for iss in row2['issn'])
                        row2_values.append(issn_string)
                    else:
                        row2_values.append(row2[head])
                sheet.append(row2_values)

def build_matches_in_chunks(wb, base_name, matches, max_rows=1_000_000):
    total = len(matches)

    if total < max_rows:
        ws = wb.create_sheet(base_name)
        write_sheet(ws, matches, True)
        apply_style(ws)
    else:
        num_parts = (total + max_rows - 1) // max_rows

        for i in range(num_parts):
            start = i * max_rows
            end = min(start + max_rows, total)
            chunk = matches[start:end]

            suffix = f"_pt{(i+1):02d}" # applies 2 digits numerical format
            ws_name = f"{base_name}{suffix}"

            ws = wb.create_sheet(ws_name)
            write_sheet(ws, chunk, True)
            apply_style(ws)

# build xlxs file inserting data and analysis, procedure
def analysis_in_xslx(ds1, ds1_unique, ds1_duplicates, ds1_name, ds2, ds2_unique, ds2_duplicates, ds2_name, in_common, matches1_2, matches2_1):
    comparison_name = f"{ds1_name}-{ds2_name}"
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, '..', 'output')
    today_date = datetime.today().strftime('%Y-%m-%d')
    file_name = today_date + '_' + comparison_name + '_Comparison.xlsx'
    file_path = os.path.join(output_dir, file_name)

    wb = Workbook() # create a new workbook

    ws_ds1 = wb.active # dataset 1 sheet
    ws_ds1.title = ds1_name 
    write_sheet(ws_ds1, ds1)
    apply_style(ws_ds1)

    ws_ds2 = wb.create_sheet(ds2_name) # dataset 2 sheet
    write_sheet(ws_ds2, ds2)
    apply_style(ws_ds2)

    ws_ds1_u_name = f"{ds1_name}_unique" # unique element sheets
    ws_ds1_u = wb.create_sheet(ws_ds1_u_name) 
    write_sheet(ws_ds1_u, ds1_unique)
    apply_style(ws_ds1_u)
    ws_ds2_u_name = f"{ds2_name}_unique"
    ws_ds2_u = wb.create_sheet(ws_ds2_u_name)
    write_sheet(ws_ds2_u, ds2_unique)
    apply_style(ws_ds2_u)

    ws_ic = wb.create_sheet('in_common') # in common elements sheet
    write_sheet(ws_ic, in_common)
    apply_style(ws_ic)

    if matches1_2 != [] and matches2_1 != []: # matches sheet
        ws_ds1_d_name = f"{ds1_name}_duplicates" # duplicates element sheets
        ws_ds1_d = wb.create_sheet(ws_ds1_d_name) 
        write_sheet(ws_ds1_d, ds1_duplicates, True)
        apply_style(ws_ds1_d)
        ws_ds2_d_name = f"{ds2_name}_duplicates"
        ws_ds2_d = wb.create_sheet(ws_ds2_d_name)
        write_sheet(ws_ds2_d, ds2_duplicates, True)
        apply_style(ws_ds2_d)

        ws_matches_name = f"matches_{ds1_name}_vs_{ds2_name}"
        build_matches_in_chunks(wb, ws_matches_name, matches1_2)
        ws_matches_name = f"matches_{ds2_name}_vs_{ds1_name}"
        build_matches_in_chunks(wb, ws_matches_name, matches2_1)

    # dataset analysis
    ws_a = wb.create_sheet('summary_analysis') # summary sheets
    ws_a['A1'] = f"Extraction date:"
    ws_a['A1'].font = Font(bold=True)
    ws_a['B1'] = datetime.today()

    ws_a['B3'] = ds1_name # dataset length table
    ws_a['C3'] = ds2_name
    ws_a['A4'] = 'Unique pubs'
    ws_a['B4'] = f"=COUNTA('{ws_ds1_u_name}'!A:A) - 1"
    ws_a['C4'] = f"=COUNTA('{ws_ds2_u_name}'!A:A) - 1"
    ws_a['A5'] = 'In common pubs'
    ws_a['B5'] = "=COUNTA(in_common!A:A) - 1"
    ws_a['C5'] = "=COUNTA(in_common!A:A) - 1"
    ws_a['A6'] = 'Total pubs'
    ws_a['B6'] = f"=COUNTA('{ds1_name}'!A:A) - 1"
    ws_a['C6'] = f"=COUNTA('{ds2_name}'!A:A) - 1"

    ws_a['G3'] = 'Unique pubs' # dataset composition percentage table
    ws_a['H3'] = 'In common pubs'
    ws_a['F4'] = ds1_name
    ws_a['G4'] = '=B4/B6'
    ws_a['H4'] = '=B5/B6'
    ws_a['F5'] = ds2_name
    ws_a['G5'] = '=C4/C6'
    ws_a['H5'] = '=C5/C6'

    for cell in ws_a['A3:H3']: #apply bold to tables headers
        for c in cell:
            c.font = Font(bold=True)
    for cell in ws_a['A3:A6']:
        for c in cell:
            c.font = Font(bold=True)
    for cell in ws_a['F3:F5']:
        for c in cell:
            c.font = Font(bold=True)
    
    for row in ws_a['G4:H5']: #apply percentage format to cells
        for cell in row:
            cell.number_format = '0.00%'

    bar = BarChart() # create Dataset Composition Clustered Column bar
    bar.type = 'col'
    bar.grouping = 'stacked'
    bar.style = 10
    bar.title = 'Unique vs In Common Publications (%)'
    bar.y_axis.title = 'Percentage'
    bar.x_axis.title = 'Category'

    bar_data = Reference(ws_a, min_col=7, max_col=8, min_row=3, max_row=5)
    bar.add_data(bar_data, titles_from_data=True)
    bar_cats = Reference(ws_a, min_col=6, min_row=4, max_row=5)
    bar.set_categories(bar_cats)

    bar.dLbls = DataLabelList() # shows labels on chart
    bar.dLbls.showVal = True
    bar.dLbls.showCatName = False
    bar.dLbls.showLegendKey = False
    bar.dLbls.showSerName = False
    bar.x_axis.delete = False # shows category labels
    bar.x_axis.tickLblPos = 'nextTo' 

    ws_a.add_chart(bar, 'A8')

    # matches analysis
    if matches1_2 != [] and matches2_1 != []:
        # sure matches table
        sure_list = [ # gets sure matches
            mat[3]
            for mat in matches1_2
            if mat[3] in ('doi', 'title; year; publisher; container_name', 'title; issn; vol_iss', 'title; year; authors')
        ]
        sure_counting = Counter(sure_list) # build a counter of confidance classes
        ws_a['B25'] = 'Matches' 
        ws_a['A26'] = 'doi'
        ws_a['B26'] = sure_counting['doi']
        ws_a['A27'] = 'Suff1'
        ws_a['B27'] = sure_counting['title; year; publisher; container_name']
        ws_a['A28'] = 'Suff2'
        ws_a['B28'] = sure_counting['title; issn; vol_iss']
        ws_a['A29'] = 'Suff3'
        ws_a['B29'] = sure_counting['title; year; authors']

        ws_a['D25'] = 'Legend' # sure matches legend
        ws_a['D26'] = 'The following field matches are considered sufficient:'
        ws_a['D27'] = 'Suff1'
        ws_a['E27'] = 'title; year; publisher; container_name'
        ws_a['D28'] = 'Suff2'
        ws_a['E28'] = 'title; issn; vol_iss'
        ws_a['D29'] = 'Suff3'
        ws_a['E29'] = 'title; year; authors'

        pie1 = PieChart() # create sure matches PieChart
        pie1.title = 'Sure matches composition'
        pie1_values = Reference(ws_a, min_col=2, min_row=26, max_row=29)
        pie1.add_data(pie1_values, titles_from_data=False)
        pie1_labels = Reference(ws_a, min_col=1, min_row=26, max_row=29)
        pie1.set_categories(pie1_labels)
        pie1.firstSliceAng = 270
        pie1.dataLabels = DataLabelList()
        pie1.dataLabels.showPercent = True
        pie1.dataLabels.showVal = False
        pie1.dLbls.showCatName = False
        pie1.dLbls.showLegendKey = False
        pie1.dLbls.showSerName = False

        ws_a.add_chart(pie1, 'A32')

        # unsure matches confidence classes table
        confidence_list = [ # gets not sure matches confidence (>= 0.75)
            mat1[2]
            for mat1 in matches1_2
            if mat1[3] not in ('doi', 'title; year; publisher; container_name', 'title; issn; vol_iss', 'title; year; authors')
            and mat1[2] >= 0.75
        ] 
        confidence_list.sort(reverse=True)
        confidence_counting = Counter(confidence_list) # build a counter of confidance classes
        ws_a['A49'] = 'Confidence'
        ws_a['B49'] = 'Number of matches'
        for i, (conf_class, conf_count) in enumerate(confidence_counting.items(), start=50): # print table
            ws_a[f"A{i}"] = conf_class
            ws_a[f"B{i}"] = conf_count
        ws_a['D49'] = 'Legend' # unsure matches confidence legend
        ws_a['D50'] = 'Number of matches with same confidence greater or equal to 0.75'
        ws_a['D51'] = 'not resulting from sure matches'

        pie2_start_row = 50 # create Piechart of unsure matches matched fields
        pie2_end_row = pie2_start_row + len(confidence_counting) - 1
        pie2 = PieChart()
        pie2.title = 'Unsure matches confidence composition'
        pie2_values = Reference(ws_a, min_col=2, min_row=pie2_start_row, max_row=pie2_end_row)
        pie2.add_data(pie2_values, titles_from_data=False)
        pie2_labels = Reference(ws_a, min_col=1, min_row=pie2_start_row, max_row=pie2_end_row)
        pie2.set_categories(pie2_labels)
        pie2.firstSliceAng = 270
        pie2.dataLabels = DataLabelList()
        pie2.dataLabels.showPercent = True
        pie2.dataLabels.showVal = False
        pie2.dLbls.showCatName = False
        pie2.dLbls.showLegendKey = False
        pie2.dLbls.showSerName = False

        ws_a.add_chart(pie2, f"A{pie2_end_row + 4}")
        
        # unsure matches fields matched classes table
        fields_list = [ # gets not sure matched field (confidence >= 1)
            mat2[3]
            for mat2 in matches1_2
            if mat2[3] not in ('doi', 'title; year; publisher; container_name', 'title; issn; vol_iss', 'title; year; authors')
            and mat2[2] >= 1
        ]
        fields_list.sort(reverse=True)
        fields_counting = Counter(fields_list) # build a counter of matchedfields classes
        table_start_row = pie2_end_row + 21
        ws_a[f"A{table_start_row}"] = 'Matched fields'
        ws_a[f"B{table_start_row }"] = 'Number of matches'
        for i, (fields_class, fields_count) in enumerate(fields_counting.items(), start=table_start_row + 1): # print table
            ws_a[f"A{i}"] = fields_class
            ws_a[f"B{i}"] = fields_count
        ws_a[f"D{table_start_row}"] = 'Legend' # unsure matches matched fields matched legend
        ws_a[f"D{table_start_row + 1}"] = 'Number of matches with same fields matched with confidence greater'
        ws_a[f"D{table_start_row + 2}"] = 'or equal to 1, not resulting from sure matches'

        pie3_start_row = table_start_row + 1 # create PieChart for unsure matched fields classes
        pie3_end_row = pie3_start_row + len(fields_counting) - 1
        pie3 = PieChart()
        pie3.title = 'Unsure matches matched fields composition'
        pie3_values = Reference(ws_a, min_col=2, min_row=pie3_start_row, max_row=pie3_end_row)
        pie3.add_data(pie3_values, titles_from_data=False)
        pie3_labels = Reference(ws_a, min_col=1, min_row=pie3_start_row, max_row=pie3_end_row)
        pie3.set_categories(pie3_labels)
        pie3.firstSliceAng = 270
        pie3.dataLabels = DataLabelList()
        pie3.dataLabels.showPercent = True
        pie3.dataLabels.showVal = False
        pie3.dLbls.showCatName = False
        pie3.dLbls.showLegendKey = False
        pie3.dLbls.showSerName = False

        ws_a.add_chart(pie3, f"A{pie3_end_row + 3}")
        
        for cell in ws_a['B25:D25']: #apply bold to tables headers
            for c in cell:
                c.font = Font(bold=True)
        for cell in ws_a['A26:A29']:
            for c in cell:
                c.font = Font(bold=True)
        for cell in ws_a['A49:D49']: #apply bold to tables headers
            for c in cell:
                c.font = Font(bold=True)
        for cell in ws_a[f"A{table_start_row}:D{table_start_row}"]: #apply bold to tables headers
            for c in cell:
                c.font = Font(bold=True)


    wb.save(file_path) # save the workbook
    print(f"[analysis] Saved {file_name} in output folder.")


#### ASKING FUNCTIONS ####
# asks user if they want to generate a report on the current comparison, procedure
def ask_for_analysis(comp_set, datasets):
    if len(comp_set) != 0:
        analysis_flag = input(f"[analysis] Do you want a report on last comparison? (y/n): ")
        if analysis_flag.lower() == 'y':
            dataset1 = datasets[0]
            dataset2 = datasets[1]
            matches1_2 = []
            matches2_1 = []
            comp_ds1_duplicates = []
            comp_ds2_duplicates = []
            if len(comp_set) == 4: # same_source_comparison
                dataset_names = re.match(r"^(.*?)\s+vs\s+(.*)$", comp_set[3])
                dataset1_name = dataset_names.group(1)
                dataset2_name = dataset_names.group(2)
                comp_ds1_unique = comp_set[1]
                comp_ds2_unique = comp_set[2]
                comp_in_common = comp_set[0]
            else: # diff_source_comparison
                dataset_names = re.match(r"^(.*?)\s+vs\s+(.*)$", comp_set[7])
                dataset1_name = dataset_names.group(1)
                dataset2_name = dataset_names.group(2)
                comp_ds1_unique = comp_set[1]
                comp_ds2_unique = comp_set[2]
                comp_ds1_duplicates = comp_set[3]
                comp_ds2_duplicates = comp_set[4]
                comp_in_common = comp_set[0]
                matches1_2 = comp_set[5]
                matches2_1 = comp_set[6]

            analysis_in_xslx(dataset1, comp_ds1_unique, comp_ds1_duplicates, dataset1_name, dataset2, comp_ds2_unique, comp_ds2_duplicates, dataset2_name, comp_in_common, matches1_2, matches2_1)